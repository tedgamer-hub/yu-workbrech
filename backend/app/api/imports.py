from __future__ import annotations

import csv
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal, cast
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.config import settings
from app.database import get_db
from app.dependencies import get_current_user
from app.models import AdmissionScore, ImportRowError, ImportTask, User

router = APIRouter()

ImportStatus = Literal["uploaded", "running", "completed", "failed"]
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}

REQUIRED_COLUMN_ALIASES = {
    "college": {"college", "school", "\u9662\u6821", "\u5b66\u6821", "\u9662\u6821\u540d\u79f0", "\u5b66\u6821\u540d\u79f0"},
    "major": {"major", "specialty", "\u4e13\u4e1a", "\u4e13\u4e1a\u540d\u79f0"},
    "province": {"province", "\u751f\u6e90\u5730", "\u7701\u4efd", "\u7701"},
    "year": {"year", "\u5e74\u4efd", "\u62db\u751f\u5e74\u4efd"},
    "min_score": {"minscore", "score", "\u6700\u4f4e\u5206", "\u5f55\u53d6\u6700\u4f4e\u5206"},
    "min_rank": {"minrank", "rank", "\u6700\u4f4e\u4f4d\u6b21", "\u5f55\u53d6\u6700\u4f4e\u4f4d\u6b21"},
}


class ImportRunRequest(BaseModel):
    strict_mode: bool = False
    max_error_samples: int = Field(default=10, ge=1, le=100)


class ImportTaskResponse(BaseModel):
    id: int
    filename: str
    storage_path: str | None
    status: str
    total_rows: int
    success_rows: int
    error_rows: int
    error_message: str | None
    created_by_user_id: int
    created_at: datetime
    started_at: datetime | None
    finished_at: datetime | None
    updated_at: datetime
    progress_percent: int
    is_terminal: bool


class ImportTaskListResponse(BaseModel):
    total: int
    items: list[ImportTaskResponse]


class ImportTaskStatusResponse(BaseModel):
    id: int
    status: str
    total_rows: int
    success_rows: int
    error_rows: int
    error_message: str | None
    started_at: datetime | None
    finished_at: datetime | None
    updated_at: datetime
    progress_percent: int
    is_terminal: bool


class ImportRowErrorResponse(BaseModel):
    id: int
    import_task_id: int
    row_number: int
    error_reason: str
    raw_row_json: str | None
    created_at: datetime


class ImportRowErrorListResponse(BaseModel):
    total: int
    items: list[ImportRowErrorResponse]


def _normalize_header(value: str) -> str:
    return value.strip().lower().replace("_", "").replace(" ", "").replace("-", "")


def _normalize_aliases() -> dict[str, set[str]]:
    return {
        canonical: {_normalize_header(alias) for alias in aliases}
        for canonical, aliases in REQUIRED_COLUMN_ALIASES.items()
    }


NORMALIZED_ALIASES = _normalize_aliases()


def _to_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    text = str(value).strip()
    if text == "":
        return None

    try:
        if "." in text:
            return int(float(text))
        return int(text)
    except ValueError:
        return None


def _resolve_import_storage_dir() -> Path:
    raw_dir = Path(settings.import_storage_dir)
    if not raw_dir.is_absolute():
        backend_dir = Path(__file__).resolve().parents[2]
        raw_dir = (backend_dir / raw_dir).resolve()

    raw_dir.mkdir(parents=True, exist_ok=True)
    return raw_dir


def _save_uploaded_file(file: UploadFile) -> Path:
    original_name = file.filename or "unnamed"
    ext = Path(original_name).suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file extension: {ext or 'none'}. Use .csv or .xlsx",
        )

    storage_dir = _resolve_import_storage_dir()
    safe_name = f"{datetime.now(UTC).strftime('%Y%m%d_%H%M%S')}_{uuid4().hex}{ext}"
    target = storage_dir / safe_name

    file.file.seek(0)
    with target.open("wb") as out:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            out.write(chunk)

    return target


def _status_progress(status_value: str) -> int:
    if status_value == "uploaded":
        return 0
    if status_value == "running":
        return 50
    return 100


def _is_terminal(status_value: str) -> bool:
    return status_value in {"completed", "failed"}


def _as_import_status(status_value: str) -> str:
    if status_value in {"uploaded", "running", "completed", "failed"}:
        return status_value
    return "failed"


def _to_response(item: ImportTask) -> ImportTaskResponse:
    status_value = _as_import_status(item.status)
    return ImportTaskResponse(
        id=item.id,
        filename=item.filename,
        storage_path=item.storage_path,
        status=status_value,
        total_rows=item.total_rows,
        success_rows=item.success_rows,
        error_rows=item.error_rows,
        error_message=item.error_message,
        created_by_user_id=item.created_by_user_id,
        created_at=item.created_at,
        started_at=item.started_at,
        finished_at=item.finished_at,
        updated_at=item.updated_at,
        progress_percent=_status_progress(status_value),
        is_terminal=_is_terminal(status_value),
    )


def _to_status_response(item: ImportTask) -> ImportTaskStatusResponse:
    status_value = _as_import_status(item.status)
    return ImportTaskStatusResponse(
        id=item.id,
        status=status_value,
        total_rows=item.total_rows,
        success_rows=item.success_rows,
        error_rows=item.error_rows,
        error_message=item.error_message,
        started_at=item.started_at,
        finished_at=item.finished_at,
        updated_at=item.updated_at,
        progress_percent=_status_progress(status_value),
        is_terminal=_is_terminal(status_value),
    )


def _to_row_error_response(item: ImportRowError) -> ImportRowErrorResponse:
    return ImportRowErrorResponse(
        id=item.id,
        import_task_id=item.import_task_id,
        row_number=item.row_number,
        error_reason=item.error_reason,
        raw_row_json=item.raw_row_json,
        created_at=item.created_at,
    )


def _check_access(item: ImportTask, current_user: User) -> None:
    if current_user.role == "admin":
        return
    if item.created_by_user_id != current_user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed to access this import task")


def _apply_visibility(stmt, current_user: User):
    if current_user.role == "admin":
        return stmt
    return stmt.where(ImportTask.created_by_user_id == current_user.id)


def _build_header_mapping(headers: list[str]) -> tuple[dict[str, str], list[str]]:
    normalized_header_to_original = {}
    for header in headers:
        if not header:
            continue
        normalized_header_to_original[_normalize_header(header)] = header

    mapping: dict[str, str] = {}
    missing: list[str] = []
    for canonical, aliases in NORMALIZED_ALIASES.items():
        found = next((normalized_header_to_original[a] for a in aliases if a in normalized_header_to_original), None)
        if found is None:
            missing.append(canonical)
        else:
            mapping[canonical] = found

    return mapping, missing


def _iter_csv_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    decode_error: UnicodeDecodeError | None = None

    for encoding in ("utf-8-sig", "gb18030", "gbk"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                if reader.fieldnames is None:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV header is required")

                headers = [str(h).strip() for h in reader.fieldnames]
                rows = []
                for row in reader:
                    rows.append({(k or "").strip(): v for k, v in row.items()})

            return headers, rows
        except UnicodeDecodeError as err:
            decode_error = err

    if decode_error is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV encoding not supported. Please use UTF-8/GBK/GB18030.",
        ) from decode_error

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="CSV parse failed")


def _iter_xlsx_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file has no active sheet")
        row_iter = sheet.iter_rows(values_only=True)
        first_row = next(row_iter, None)
        if first_row is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file is empty")

        headers = [str(v).strip() if v is not None else "" for v in first_row]
        rows = []
        for values in row_iter:
            row_obj: dict[str, object] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                cell_value = values[idx] if idx < len(values) else None
                row_obj[header] = cell_value
            rows.append(row_obj)
    finally:
        workbook.close()

    return headers, rows


def _load_rows(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _iter_csv_rows(path)
    if ext == ".xlsx":
        return _iter_xlsx_rows(path)

    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file type")


def _parse_row(row: dict[str, object], mapping: dict[str, str]) -> tuple[dict[str, Any] | None, str | None]:
    college = str(row.get(mapping["college"], "")).strip()
    major = str(row.get(mapping["major"], "")).strip()
    province = str(row.get(mapping["province"], "")).strip()

    year = _to_int(row.get(mapping["year"]))
    min_score = _to_int(row.get(mapping["min_score"]))
    min_rank = _to_int(row.get(mapping["min_rank"]))

    if not college:
        return None, "college is empty"
    if not major:
        return None, "major is empty"
    if not province:
        return None, "province is empty"
    if year is None or year < 2000 or year > 2100:
        return None, "year is invalid"
    if min_score is None or min_score < 0 or min_score > 750:
        return None, "min_score is invalid"
    if min_rank is None or min_rank < 0:
        return None, "min_rank is invalid"

    return {
        "college_name": college,
        "major_name": major,
        "province": province,
        "year": year,
        "min_score": min_score,
        "min_rank": min_rank,
    }, None


def _upsert_admission_score(
    db: Session,
    *,
    task: ImportTask,
    source_row_number: int,
    parsed: dict[str, Any],
) -> None:
    db.flush()
    existing = db.scalar(
        select(AdmissionScore).where(
            AdmissionScore.college_name == parsed["college_name"],
            AdmissionScore.major_name == parsed["major_name"],
            AdmissionScore.province == parsed["province"],
            AdmissionScore.year == parsed["year"],
        )
    )

    if existing is None:
        db.add(
            AdmissionScore(
                college_name=parsed["college_name"],
                major_name=parsed["major_name"],
                province=parsed["province"],
                year=parsed["year"],
                min_score=parsed["min_score"],
                min_rank=parsed["min_rank"],
                source_filename=task.filename,
                source_row_number=source_row_number,
                import_task_id=task.id,
            )
        )
        return

    existing.min_score = parsed["min_score"]
    existing.min_rank = parsed["min_rank"]
    existing.source_filename = task.filename
    existing.source_row_number = source_row_number
    existing.import_task_id = task.id


def _execute_import(
    task: ImportTask,
    *,
    db: Session,
    strict_mode: bool,
    max_error_samples: int,
) -> None:
    if not task.storage_path:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Task has no file path")

    file_path = Path(task.storage_path)
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file not found")

    headers, rows = _load_rows(file_path)
    mapping, missing_columns = _build_header_mapping(headers)
    if missing_columns:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing required columns: {', '.join(missing_columns)}",
        )

    db.execute(delete(ImportRowError).where(ImportRowError.import_task_id == task.id))
    db.execute(delete(AdmissionScore).where(AdmissionScore.import_task_id == task.id))

    total = 0
    success = 0
    errors = 0
    error_samples: list[str] = []

    for row_number, row in enumerate(rows, start=2):
        total += 1
        parsed, parse_error = _parse_row(row, mapping)
        if parse_error:
            errors += 1
            if len(error_samples) < max_error_samples:
                error_samples.append(f"row {row_number}: {parse_error}")

            db.add(
                ImportRowError(
                    import_task_id=task.id,
                    row_number=row_number,
                    error_reason=parse_error,
                    raw_row_json=json.dumps(row, ensure_ascii=False, separators=(",", ":")),
                )
            )
            if strict_mode:
                break
            continue

        if parsed is None:
            errors += 1
            reason = "row parse returned empty result"
            if len(error_samples) < max_error_samples:
                error_samples.append(f"row {row_number}: {reason}")
            db.add(
                ImportRowError(
                    import_task_id=task.id,
                    row_number=row_number,
                    error_reason=reason,
                    raw_row_json=json.dumps(row, ensure_ascii=False, separators=(",", ":")),
                )
            )
            if strict_mode:
                break
            continue

        _upsert_admission_score(
            db,
            task=task,
            source_row_number=row_number,
            parsed=cast(dict[str, Any], parsed),
        )
        success += 1

    task.total_rows = total
    task.success_rows = success
    task.error_rows = errors

    if total == 0:
        task.status = "failed"
        task.error_message = "No data rows found"
        return

    if strict_mode and errors > 0:
        task.status = "failed"
    elif success == 0:
        task.status = "failed"
    else:
        task.status = "completed"

    task.error_message = "; ".join(error_samples) if error_samples else None


@router.post("/upload", response_model=ImportTaskResponse, status_code=status.HTTP_201_CREATED)
def upload_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportTaskResponse:
    saved_file = _save_uploaded_file(file)

    task = ImportTask(
        filename=file.filename or saved_file.name,
        storage_path=str(saved_file),
        status="uploaded",
        created_by_user_id=current_user.id,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return _to_response(task)


@router.get("", response_model=ImportTaskListResponse)
def list_imports(
    status_filter: ImportStatus | None = Query(default=None, alias="status"),
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportTaskListResponse:
    base_stmt = _apply_visibility(select(ImportTask), current_user)
    count_stmt = _apply_visibility(select(func.count()).select_from(ImportTask), current_user)

    if status_filter:
        base_stmt = base_stmt.where(ImportTask.status == status_filter)
        count_stmt = count_stmt.where(ImportTask.status == status_filter)

    total = db.scalar(count_stmt) or 0
    items = db.scalars(base_stmt.order_by(ImportTask.id.desc()).offset(offset).limit(limit)).all()
    return ImportTaskListResponse(total=total, items=[_to_response(item) for item in items])


@router.get("/{import_id}", response_model=ImportTaskResponse)
def get_import_task(
    import_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportTaskResponse:
    task = db.get(ImportTask, import_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import task not found")

    _check_access(task, current_user)
    return _to_response(task)


@router.get("/{import_id}/status", response_model=ImportTaskStatusResponse)
def get_import_status(
    import_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportTaskStatusResponse:
    task = db.get(ImportTask, import_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import task not found")

    _check_access(task, current_user)
    return _to_status_response(task)


@router.get("/{import_id}/errors", response_model=ImportRowErrorListResponse)
def list_import_errors(
    import_id: int,
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportRowErrorListResponse:
    task = db.get(ImportTask, import_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import task not found")

    _check_access(task, current_user)

    count_stmt = select(func.count()).select_from(ImportRowError).where(ImportRowError.import_task_id == import_id)
    list_stmt = (
        select(ImportRowError)
        .where(ImportRowError.import_task_id == import_id)
        .order_by(ImportRowError.row_number.asc())
        .offset(offset)
        .limit(limit)
    )

    total = db.scalar(count_stmt) or 0
    items = db.scalars(list_stmt).all()
    return ImportRowErrorListResponse(total=total, items=[_to_row_error_response(item) for item in items])


@router.post("/{import_id}/run", response_model=ImportTaskResponse)
def run_import(
    import_id: int,
    payload: ImportRunRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ImportTaskResponse:
    task = db.get(ImportTask, import_id)
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import task not found")

    _check_access(task, current_user)

    if task.status == "running":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Import task is already running")

    task.status = "running"
    task.total_rows = 0
    task.success_rows = 0
    task.error_rows = 0
    task.error_message = None
    task.started_at = datetime.now(UTC)
    task.finished_at = None
    db.commit()
    db.refresh(task)

    try:
        _execute_import(task, db=db, strict_mode=payload.strict_mode, max_error_samples=payload.max_error_samples)
    except HTTPException as err:
        db.rollback()
        task.status = "failed"
        task.error_message = str(err.detail)
    except Exception:
        db.rollback()
        task.status = "failed"
        task.error_message = "Unexpected import error"

    task.finished_at = datetime.now(UTC)
    db.commit()
    db.refresh(task)
    return _to_response(task)
